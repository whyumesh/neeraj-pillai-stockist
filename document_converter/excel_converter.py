"""
Excel converter - Convert extracted data to Excel format
"""
from pathlib import Path
from typing import Dict, List, Any, Optional
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

logger = logging.getLogger(__name__)


class ExcelConverter:
    """Convert extracted data to Excel format"""
    
    def __init__(self, config=None):
        """Initialize Excel converter"""
        from .config_loader import ConfigLoader
        self.config = config or ConfigLoader()
        
        # Excel formatting styles
        self.header_fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid"
        )
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def convert_po_to_excel(self,
                           data: Dict[str, Any],
                           output_path: Path,
                           source_filename: str) -> bool:
        """
        Convert Purchase Order data to Excel - output as-is, no Summary/Metadata sheets
        Uses adaptive headers from extracted items (same approach as Stock & Sales)
        
        Args:
            data: Extracted PO data
            output_path: Output Excel file path
            source_filename: Source file name for metadata
            
        Returns:
            True if successful
        """
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            items = data.get("items", [])
            
            if not items:
                logger.warning(f"No items to write for PO: {source_filename}")
                # Create an empty sheet if no data
                wb.create_sheet("Sheet1")
            else:
                ws = wb.create_sheet("Sheet1")
                # Use the same adaptive method as Stock & Sales
                self._add_items_to_sheet(ws, items)
            
            # Save
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            logger.info(f"Created PO Excel file: {output_path} with {len(items)} items")
            return True
        
        except Exception as e:
            logger.error(f"Error creating PO Excel file: {e}", exc_info=True)
            return False
    
    def convert_stock_sales_to_excel(self,
                                    data: Dict[str, Any],
                                    output_path: Path,
                                    source_filename: str) -> bool:
        """
        Convert Stock & Sales data to Excel - output as-is, no Summary/Metadata sheets
        
        Args:
            data: Extracted stock & sales data
            output_path: Output Excel file path
            source_filename: Source file name for metadata
            
        Returns:
            True if successful
        """
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            items = data.get("items", [])
            sections = data.get("sections", [])
            
            if not items:
                logger.warning(f"No items to write for {source_filename}")
                wb.create_sheet("Sheet1")
                wb.save(output_path)
                return True
            
            # Group items by section
            items_by_section = {}
            for item in items:
                section = item.get("section", "UNSPECIFIED")
                if section not in items_by_section:
                    items_by_section[section] = []
                items_by_section[section].append(item)
            
            # If there are multiple sections, create one sheet per section
            if len(items_by_section) > 1:
                for section, section_items in items_by_section.items():
                    if section_items:
                        # Excel sheet name limit is 31 chars
                        sheet_name = section[:31] if section != "UNSPECIFIED" else "Sheet1"
                        ws_section = wb.create_sheet(sheet_name)
                        self._add_items_to_sheet(ws_section, section_items)
            # Single section or no sections - create one sheet
            else:
                ws = wb.create_sheet("Sheet1")
                self._add_items_to_sheet(ws, items)
            
            # Save
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            logger.info(f"Created Stock & Sales Excel file: {output_path} with {len(items)} items")
            return True
        
        except Exception as e:
            logger.error(f"Error creating Stock & Sales Excel file: {e}", exc_info=True)
            return False
    
    def convert_table_data_to_excel(self,
                                   headers: List[str],
                                   rows: List[List[Any]],
                                   output_path: Path,
                                   source_filename: str,
                                   sheet_name: str = "Data") -> bool:
        """
        Convert generic table data to Excel
        
        Args:
            headers: Column headers
            rows: Data rows
            output_path: Output Excel file path
            source_filename: Source file name
            sheet_name: Sheet name
            
        Returns:
            True if successful
        """
        try:
            # Create DataFrame
            df = pd.DataFrame(rows, columns=headers)
            
            # Write to Excel
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get worksheet for formatting
                worksheet = writer.sheets[sheet_name]
                
                # Format headers
                for cell in worksheet[1]:
                    cell.fill = self.header_fill
                    cell.font = self.header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = self.border
                
                # Format data rows
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.border = self.border
                
                # Auto-adjust widths
                self._auto_adjust_widths(worksheet)
            
            logger.info(f"Created Excel file from table data: {output_path}")
            return True
        
        except Exception as e:
            logger.error(f"Error creating Excel file: {e}")
            return False
    
    def _add_items_to_sheet(self, worksheet, items: List[Dict[str, Any]]):
        """
        Add items to worksheet with formatting - PRESERVE EXACT COLUMN ORDER AND HEADER NAMES
        
        This method preserves the exact structure from the PDF, including:
        - Original header names (including multi-row headers like "Opening Qty - Qty")
        - Exact column order
        - All columns as they appear in source
        
        Args:
            worksheet: OpenPyXL worksheet object
            items: List of item dictionaries
        """
        if not items:
            logger.warning("No items to add to worksheet")
            return
        
        # PRESERVE EXACT COLUMN ORDER: Use keys in order they appear in first item
        # Don't filter or rearrange - preserve exact structure
        
        if not items:
            return
        
        # Get all unique keys in order of first appearance across all items
        # This preserves the column order from extraction
        header_order = []
        seen_keys = set()
        
        # Collect keys in order from first item (this determines column order)
        for key in items[0].keys():
            if key not in seen_keys:
                header_order.append(key)
                seen_keys.add(key)
        
        # Also check other items for any missing keys
        for item in items[1:]:
            for key in item.keys():
                if key not in seen_keys:
                    header_order.append(key)
                    seen_keys.add(key)
        
        # Use headers from items directly - preserve exact column order from PDF
        # This works for both PO (Sr, Product Name, Pack, Qty, Free, Rate, Amount) 
        # and Stock & Sales (Item Description, Opening Qty, Opening Value, etc.)
        headers = header_order.copy()
        
        # Add "section" at the end if it exists (but only once)
        seen_normalized = set()
        normalize_col_name = lambda name: name.lower().replace('_', ' ').replace('-', ' ').strip()
        
        for field in ["section", "Section"]:
            if field in header_order and normalize_col_name(field) not in seen_normalized:
                if field in headers:
                    headers.remove(field)  # Remove from current position
                headers.append(field)  # Add at end
                seen_normalized.add(normalize_col_name(field))
                break
        
        if not headers:
            logger.error("No headers found - cannot write Excel file")
            return
        
        logger.info(f"Writing {len(items)} items with {len(headers)} columns: {headers}")
        
        # Write headers (use original header names)
        worksheet.append(headers)
        
        # Format headers
        for cell in worksheet[worksheet.max_row]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.border
        
        # Add items - preserve exact column order and use standard header names
        for item_idx, item in enumerate(items):
            row = []
            for header in headers:
                # Try to get value by exact header name first
                value = item.get(header)
                
                # If not found, try normalized version (case-insensitive, handle variations)
                if value is None:
                    # Try case-insensitive match
                    header_lower = header.lower()
                    for key in item.keys():
                        if key.lower() == header_lower:
                            value = item.get(key)
                            break
                
                # If still not found, try safe_key version (for backward compatibility)
                if value is None:
                    safe_key = header.lower().replace(' ', '_').replace('-', '_').replace('/', '_')
                    safe_key = re.sub(r'[^a-z0-9_]', '', safe_key)
                    value = item.get(safe_key)
                
                # Handle None values
                if value is None:
                    value = ""
                
                row.append(value)
            worksheet.append(row)
        
        # Format data rows
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.border = self.border
                # Right-align numeric columns
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
        
        # Auto-adjust widths
        self._auto_adjust_widths(worksheet)
        
        logger.info(f"Added {len(items)} rows to worksheet with {len(headers)} columns")
    
    def _auto_adjust_widths(self, worksheet):
        """Auto-adjust column widths"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            
            adjusted_width = min(max_length + 3, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

